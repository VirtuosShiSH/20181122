import os
import re
import shutil
import openpyxl
import webbrowser
import tkinter.filedialog
import requests
import time
#wb = openpyxl.load_workbook('c:\\Users\\shiso\\Documents\\Repo\\20181122\\PythonScript\\PUBG_QA\\PUBG_QA.XLSX')
#cwd = os.getcwd()
#wb = openpyxl.load_workbook(cwd+'\PUBG_QA\PUBG_QA.XLSX', data_only = 1)
wb =openpyxl.load_workbook(tkinter.filedialog.askopenfilename(title='Choose Excel File...'), data_only = 1)
sheet = wb['Issues']
row_count = sheet.max_row
KeyDic = {}
for i in range(2,row_count+1):
    myKey = sheet.cell(row = i, column = 3).value
    jiraKey = sheet.cell(row = i, column = 4).value
    if (myKey!=None) & (jiraKey!=None):
        KeyDic[myKey] = jiraKey
        #print(str(myKey) + '  ' + str(jiraKey))
#picDir = 'C:\\Users\\shiso\\Documents\\Virtuos\\PUBG_QA\\test_files'
picDir = tkinter.filedialog.askdirectory(title='Choose Pic Folder...')
oldFileNameList = os.listdir(picDir)
LinksToOpen = []
myUniqueLinks = []
for name in oldFileNameList:
    separator = ''
    if '_' in name:
        separator = '_'
    else:
        if '.' in name:
            separator = '.'
    if separator != '':
        fName = name.split(separator)[0]
        fExt =  name.split(separator)[1]
        for key in KeyDic.keys():
            if fName == str(key):
                newName = 'PUBG_'+KeyDic[key]+separator+fExt
                print(str(key)+' >> '+newName)
                os.chdir(picDir)
                LinksToOpen.append("https://jira.krafton.com/browse/PUBG-"+KeyDic[key])
                #shutil.move(name,newName)                       #Rename and replace
                #newName = picDir+'\\NEW\\'+newName             #Rename + Copy
                #shutil.copyfile(name,newName)
                mySet = set(LinksToOpen)
                myUniqueLinks = list(mySet)
for link in myUniqueLinks:
    print(link)
    time.sleep(0.2)
    webbrowser.open_new_tab(link)
print("Done")